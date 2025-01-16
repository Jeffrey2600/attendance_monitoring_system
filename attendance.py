import face_recognition
import cv2
import pickle
import pandas as pd
from datetime import datetime
import openpyxl

def mark_attendance(video_source=0, encoding_file="face_encodings.pkl", excel_file="attendance.xlsx"):
    # Load face encodings
    with open(encoding_file, "rb") as f:
        known_faces = pickle.load(f)

    # Try to load the Excel sheet, otherwise create a new one
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        print(f"Existing data loaded from {excel_file}")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Roll Number", "Name", "Timestamp", "Accuracy"])  # Adding column headers
        print(f"No existing data found. A new file will be created at {excel_file}")

    # Start video capture
    cap = cv2.VideoCapture(video_source)
    print("Press 'q' to quit.")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            for roll_no, data in known_faces.items():
                matches = face_recognition.compare_faces([data["encoding"]], face_encoding, tolerance=0.5)

                if any(matches):
                    name = data["name"]
                    distance = face_recognition.face_distance([data["encoding"]], face_encoding)[0]
                    accuracy = round((1 - distance) * 100, 2)

                    if accuracy >= 70:  # Accuracy threshold check
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                        # Check if already marked for the day
                        already_marked = False
                        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                            if row[0] == roll_no:
                                already_marked = True
                                break

                        if not already_marked:
                            sheet.append([roll_no, name, timestamp, accuracy])
                            print(f"Attendance marked for {name} ({roll_no}) with accuracy {accuracy}%")

                        # Display rectangle and name with accuracy
                        top, right, bottom, left = face_location
                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                        cv2.putText(frame, f"{name} ({accuracy}%)", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                    else:
                        print(f"Skipped {name} ({roll_no}) due to low accuracy: {accuracy}%")

        cv2.imshow("Face Detection", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

    wb.save(excel_file)
    print(f"Attendance saved to {excel_file}.")

# Run the attendance marking system
mark_attendance()
